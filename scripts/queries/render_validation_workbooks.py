"""Render per-reviewer validation workbooks from the campaign manifest.

Input:  validation_campaign.json (from export_validation_claims.py, run on prod)
Output: one .xlsx per reviewer with three sheets —
          Calibration  (identical across reviewers; judged jointly first)
          Reliability  (identical across reviewers; independent, no discussion)
          Assigned     (this reviewer's bulk papers)
        plus answer_key.json (claim -> config -> DU mapping). The answer key is
        for the ANALYSIS phase only — do not share it with reviewers.

Workbooks contain NO config information; every link carries ?blind=1.
REFUSES to overwrite existing workbooks (reviewers annotate them in place —
never regenerate; make surgical edits instead).

Usage:
  uv run python scripts/queries/render_validation_workbooks.py \
      /path/to/validation_campaign.json /path/to/output_dir \
      --reviewers "Anthony,Aidan,Brian"
"""
import argparse
import json
import os
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

BASE_URL = os.getenv("PDL_BASE_URL", "http://localhost:8000")
HEADERS = ["claim_key", "bibcode", "paper_title", "instrument", "observatory",
           "window_start", "window_end", "validation_link", "reviewed (x)", "notes"]
WIDTHS = [14, 22, 50, 16, 14, 22, 22, 46, 12, 40]


def claim_row(c):
    url = f"{BASE_URL}/validate/{c['review_du']}?blind=1"
    return [c["claim_key"], c["bibcode"], c["paper_title"][:120], c["instrument"],
            c["observatory"], c["window_start"] or "", c["window_end"] or "",
            url, "", ""]


def add_sheet(wb, title, claim_list, note):
    ws = wb.create_sheet(title)
    ws.append([note])
    ws["A1"].font = Font(italic=True, color="666666")
    ws.append(HEADERS)
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDF4FF")
    for c in sorted(claim_list, key=lambda x: (x["bibcode"], x["claim_key"])):
        ws.append(claim_row(c))
        link_cell = ws.cell(row=ws.max_row, column=8)
        link_cell.hyperlink = link_cell.value
        link_cell.font = Font(color="0563C1", underline="single")
    for i, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("manifest")
    ap.add_argument("outdir")
    ap.add_argument("--reviewers", default=None,
                    help="Comma-separated display names replacing reviewer_1..3")
    args = ap.parse_args()

    m = json.load(open(args.manifest))
    meta, claims = m["meta"], m["claims"]
    by_key = {c["claim_key"]: c for c in claims}
    by_paper = {}
    for c in claims:
        by_paper.setdefault(c["paper_id"], []).append(c)

    names = (args.reviewers.split(",") if args.reviewers else meta["reviewers"])
    assert len(names) == len(meta["reviewers"]), "need exactly %d reviewer names" % len(meta["reviewers"])
    rename = dict(zip(meta["reviewers"], [n.strip() for n in names]))

    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)

    calibration = [by_key[k] for k in meta["calibration_keys"]]
    reliability = [c for p in meta["reliability_papers"] for c in by_paper.get(p, [])]

    for internal, display in rename.items():
        path = outdir / f"validation_{display.lower().replace(' ', '_')}.xlsx"
        if path.exists():
            sys.exit(f"REFUSING to overwrite {path} — reviewers may have annotated it. "
                     "Move it aside deliberately if you truly want to regenerate.")
        bulk_papers = [p for p, r in meta["assignment"].items() if r == internal]
        bulk = [c for p in bulk_papers for c in by_paper.get(p, [])]

        wb = Workbook()
        wb.remove(wb.active)
        add_sheet(wb, "Calibration", calibration,
                  "Round 1 — judged by ALL reviewers, then discussed jointly to freeze the rubric. "
                  "Do this before anything else.")
        add_sheet(wb, "Reliability", reliability,
                  "Reviewed by ALL reviewers INDEPENDENTLY (no discussion) — this measures "
                  "inter-rater agreement. Judge in the linked page; mark 'x' here when done.")
        add_sheet(wb, "Assigned", bulk,
                  f"Your bulk assignment ({len(bulk_papers)} papers, {len(bulk)} claims). "
                  "Judge in the linked page (approve/reject/needs-review + notes); mark 'x' here when done.")
        wb.save(path)
        print(f"{display}: calibration={len(calibration)} reliability={len(reliability)} "
              f"bulk={len(bulk)} -> {path}")

    key_path = outdir / "answer_key.json"
    if not key_path.exists():
        json.dump({c["claim_key"]: c["dus"] for c in claims}, open(key_path, "w"), indent=1)
        print(f"answer key (DO NOT SHARE): {key_path}")


if __name__ == "__main__":
    main()
