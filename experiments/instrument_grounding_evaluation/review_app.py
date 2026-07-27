#!/usr/bin/env python3
"""One-at-a-time review UI for catalog_v2_review_queue.xlsx.

Run:  uv run python review_app.py   →  open http://localhost:8765
Verdicts auto-save to review_verdicts.json on every click (safe to stop/resume;
the xlsx itself is never written, so it can stay open in Excel).

Keys: 1 = ship proposed · 2 = keep current · 3 = edit (focus comment box) ·
←/→ = prev/next · Enter (in comment box) = save+next
"""
import json
from http.server import BaseHTTPRequestHandler, HTTPServer
from pathlib import Path

from openpyxl import load_workbook

HERE = Path(__file__).parent
XLSX = HERE / "catalog_v2_review_queue.xlsx"
VERDICTS = HERE / "review_verdicts.json"
PORT = 8765

ws = load_workbook(XLSX, read_only=True)["review_queue"]
HEADER = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

import re  # noqa: E402

SPASE_RAW = HERE / "spase_raw_data.jsonl"
SPASE = {}
if SPASE_RAW.exists():
    for line in SPASE_RAW.read_text().splitlines():
        if line.strip():
            r = json.loads(line)
            SPASE[r["instrument_code"]] = r

ITEMS = []
for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
    d = dict(zip(HEADER, row))
    d["_idx"] = i
    # find a spase URI in the id (handles plain ids and merge-row "(mission, instrument)" tuples)
    uris = re.findall(r"spase://[\w/.\-]+", d.get("id") or "")
    inst_uri = next((u for u in uris if "/Instrument/" in u), None)
    any_uri = inst_uri or (uris[0] if uris else None)
    if any_uri:
        d["spase_url"] = "https://spase-metadata.org/" + any_uri.removeprefix("spase://") + ".html"
    rec = SPASE.get(inst_uri)
    if rec:
        d["spase_full"] = (
            f"ResourceName: {rec.get('spase_resource_name') or ''}\n"
            f"InstrumentType: {rec.get('spase_instrument_type') or ''}\n\n"
            f"{rec.get('spase_description') or '(no description in SPASE record)'}"
        )
    ITEMS.append(d)

PAGE = """<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>Catalog v2 review</title>
<style>
 body{font-family:-apple-system,Helvetica,sans-serif;margin:0;background:#f4f5f7;color:#1b1f24}
 header{display:flex;align-items:center;gap:14px;padding:10px 18px;background:#fff;border-bottom:1px solid #ddd;position:sticky;top:0}
 #progressbar{flex:1;height:8px;background:#e3e5e8;border-radius:4px;overflow:hidden}
 #progressfill{height:100%;background:#2da44e;width:0}
 select{padding:4px}
 .card{max-width:1100px;margin:18px auto;background:#fff;border:1px solid #ddd;border-radius:10px;padding:20px 24px;box-shadow:0 1px 3px rgba(0,0,0,.06)}
 .badge{display:inline-block;padding:2px 10px;border-radius:10px;font-size:12px;font-weight:600;color:#fff;margin-right:8px}
 .t1{background:#cf222e}.t2{background:#bf8700}.t3{background:#57606a}
 .type{color:#57606a;font-size:13px}
 h2{margin:6px 0 2px;font-size:17px;word-break:break-all}
 .note{background:#fff8c5;border:1px solid #d4a72c66;border-radius:6px;padding:10px 12px;margin:12px 0;font-size:14px}
 .cols{display:flex;gap:14px}
 .col{flex:1;border:1px solid #e3e5e8;border-radius:6px;padding:10px 12px;font-size:13.5px;line-height:1.45;white-space:pre-wrap}
 .col h4{margin:0 0 6px;font-size:12px;text-transform:uppercase;letter-spacing:.04em;color:#57606a}
 .proposed{border-color:#2da44e88;background:#f0fff4}
 .btns{display:flex;gap:10px;margin-top:16px;align-items:center;flex-wrap:wrap}
 button{padding:9px 16px;border-radius:6px;border:1px solid #ccc;background:#fff;font-size:14px;cursor:pointer}
 button:hover{background:#f3f4f6}
 .ok{background:#2da44e;border-color:#2da44e;color:#fff}.ok:hover{background:#2c974b}
 .keep{background:#0969da;border-color:#0969da;color:#fff}.keep:hover{background:#0a5fc2}
 .edit{background:#bf8700;border-color:#bf8700;color:#fff}.edit:hover{background:#a87b00}
 #comment{flex:1;min-width:280px;padding:8px;border:1px solid #ccc;border-radius:6px;font-size:14px}
 .nav{color:#57606a;font-size:13px;margin-left:auto;white-space:nowrap}
 .navbtn{padding:6px 10px}
 .done{color:#2da44e;font-weight:600}
 kbd{background:#eee;border-radius:3px;padding:1px 5px;font-size:11px;border:1px solid #ccc}
 .hint{color:#8b949e;font-size:12px;margin-top:10px}
</style></head><body>
<header>
 <select id="tierfilter" onchange="applyFilter()">
   <option value="all">All tiers</option><option value="1" selected>Tier 1 — decisions</option>
   <option value="2">Tier 2 — spot-check</option><option value="3">Tier 3 — hedges</option>
 </select>
 <select id="statefilter" onchange="applyFilter()">
   <option value="todo" selected>Unreviewed</option><option value="all">All</option><option value="done">Reviewed</option>
 </select>
 <div id="progressbar"><div id="progressfill"></div></div>
 <div class="nav" id="counter"></div>
</header>
<div id="card"></div>
<script>
let items=[],verdicts={},view=[],pos=0;
async function init(){
  items=await (await fetch('/api/items')).json();
  verdicts=await (await fetch('/api/verdicts')).json();
  applyFilter();
}
function applyFilter(){
  const t=document.getElementById('tierfilter').value, s=document.getElementById('statefilter').value;
  view=items.filter(it=>(t==='all'||String(it.tier)===t))
            .filter(it=>{const v=verdicts[it._idx];return s==='all'||(s==='done'?!!v:!v);});
  pos=0; render();
}
function esc(s){return (s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;');}
function render(){
  const total=items.filter(i=>String(i.tier)===document.getElementById('tierfilter').value||document.getElementById('tierfilter').value==='all');
  const done=total.filter(i=>verdicts[i._idx]).length;
  document.getElementById('progressfill').style.width=total.length?100*done/total.length+'%':'0';
  document.getElementById('counter').textContent=`${done}/${total.length} reviewed · showing ${view.length?pos+1:0}/${view.length}`;
  const c=document.getElementById('card');
  if(!view.length){c.innerHTML='<div class="card"><h2 class="done">Nothing left in this filter 🎉</h2><p>Switch tier or state filter above.</p></div>';return;}
  const it=view[pos], v=verdicts[it._idx]||{};
  c.innerHTML=`<div class="card">
    <div><span class="badge t${it.tier}">TIER ${it.tier}</span><span class="type">${esc(it.type)}</span></div>
    <h2>${esc(it.name||'')} <span style="color:#8b949e;font-weight:400">${esc(it.id)}</span>
      ${it.spase_url?` <a href="${it.spase_url}" target="_blank" style="font-size:13px;font-weight:400">SPASE record ↗</a>`:''}</h2>
    ${it.review_note?`<div class="note">📝 ${esc(it.review_note)}</div>`:''}
    <div class="cols">
      <div class="col"><h4>Current / prod</h4>${esc(it.current_text)||'<i>(empty)</i>'}</div>
      <div class="col proposed"><h4>Proposed</h4>${esc(it.proposed_text)}</div>
    </div>
    ${it.spase_full?`<details style="margin-top:12px"><summary style="cursor:pointer;color:#0969da;font-size:13.5px">Full SPASE record (cached ground truth)</summary><div class="col" style="margin-top:8px;background:#f6f8fa">${esc(it.spase_full)}</div></details>`:''}
    <div class="btns">
      <button class="ok" onclick="decide('ok')">✓ Ship proposed <kbd>1</kbd></button>
      <button class="keep" onclick="decide('keep')">Keep current <kbd>2</kbd></button>
      <button class="edit" onclick="document.getElementById('comment').focus()">✎ Edit… <kbd>3</kbd></button>
      <input id="comment" placeholder="comment / edited text (Enter = save as 'edit' + next)" value="${esc(v.comments||'')}">
      <span class="nav"><button class="navbtn" onclick="move(-1)">←</button> <button class="navbtn" onclick="move(1)">→</button></span>
    </div>
    ${v.verdict?`<p class="done">Saved: ${v.verdict}${v.comments?' — '+esc(v.comments):''}</p>`:''}
    <p class="hint">Keys: <kbd>1</kbd> ship · <kbd>2</kbd> keep · <kbd>3</kbd> edit · <kbd>←</kbd><kbd>→</kbd> navigate. Auto-saves to review_verdicts.json.</p>
  </div>`;
  document.getElementById('comment').addEventListener('keydown',e=>{if(e.key==='Enter'){decide('edit');}e.stopPropagation();});
}
async function decide(verdict){
  const it=view[pos], comments=document.getElementById('comment').value;
  verdicts[it._idx]={verdict,comments};
  await fetch('/api/verdict',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({idx:it._idx,id:it.id,type:it.type,verdict,comments})});
  if(document.getElementById('statefilter').value==='todo'){view.splice(pos,1);if(pos>=view.length)pos=Math.max(0,view.length-1);}
  else move(1,false);
  render();
}
function move(d,rerender=true){pos=Math.min(Math.max(0,pos+d),Math.max(0,view.length-1));if(rerender)render();}
document.addEventListener('keydown',e=>{
  if(e.target.id==='comment')return;
  if(e.key==='1')decide('ok');else if(e.key==='2')decide('keep');
  else if(e.key==='3')document.getElementById('comment').focus();
  else if(e.key==='ArrowLeft')move(-1);else if(e.key==='ArrowRight')move(1);
});
init();
</script></body></html>"""


def load_verdicts():
    return json.loads(VERDICTS.read_text()) if VERDICTS.exists() else {}


class H(BaseHTTPRequestHandler):
    def _send(self, body, ctype="application/json"):
        data = body if isinstance(body, bytes) else json.dumps(body).encode()
        self.send_response(200)
        self.send_header("Content-Type", ctype)
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def do_GET(self):
        if self.path == "/":
            self._send(PAGE.encode(), "text/html; charset=utf-8")
        elif self.path == "/api/items":
            self._send(ITEMS)
        elif self.path == "/api/verdicts":
            self._send(load_verdicts())
        else:
            self.send_error(404)

    def do_POST(self):
        if self.path == "/api/verdict":
            n = int(self.headers["Content-Length"])
            payload = json.loads(self.rfile.read(n))
            v = load_verdicts()
            v[str(payload["idx"])] = payload
            VERDICTS.write_text(json.dumps(v, indent=1))
            self._send({"ok": True})
        else:
            self.send_error(404)

    def log_message(self, *a):
        pass


if __name__ == "__main__":
    print(f"{len(ITEMS)} items loaded. Open http://localhost:{PORT}")
    HTTPServer(("127.0.0.1", PORT), H).serve_forever()
